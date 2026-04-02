"""
PDF to Excel Converter
Extracts tables from PDF to XLSX using pdfplumber
"""
import os
import logging
from typing import Dict, Any
import pdfplumber

from app.core.conversion.converters.base import BaseConverter

logger = logging.getLogger(__name__)


class PDFToExcelConverter(BaseConverter):
    """Extract tables from PDF to Excel"""

    async def convert(
        self,
        input_path: str,
        output_dir: str,
        output_file_id: str,
        options: Dict[str, Any]
    ) -> str:
        """
        Extract tables from PDF to Excel

        Options:
            - pages: List of page numbers (1-indexed). None = all pages.
        """
        self._validate_input(input_path, [".pdf"])

        output_path = self._get_output_path(output_dir, output_file_id, ".xlsx")

        try:
            logger.info(f"Extracting tables from PDF to Excel: {input_path}")

            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Extracted Tables"

            target_pages = options.get("pages", None)
            row_offset = 1
            tables_found = 0

            with pdfplumber.open(input_path) as pdf:
                for i, page in enumerate(pdf.pages):
                    if target_pages and (i + 1) not in target_pages:
                        continue

                    tables = page.extract_tables()

                    for table in tables:
                        tables_found += 1
                        # Write header row with page info
                        ws.cell(row=row_offset, column=1, value=f"--- Table from Page {i + 1} ---")
                        row_offset += 1

                        for row_data in table:
                            for col_idx, cell_value in enumerate(row_data, 1):
                                ws.cell(row=row_offset, column=col_idx, value=cell_value)
                            row_offset += 1
                        row_offset += 1  # Empty row between tables

            if tables_found == 0:
                # No tables found, extract all text as fallback
                ws.cell(row=1, column=1, value="No tables found. Raw text extracted:")
                row_offset = 3
                with pdfplumber.open(input_path) as pdf:
                    for i, page in enumerate(pdf.pages):
                        text = page.extract_text()
                        if text:
                            ws.cell(row=row_offset, column=1, value=f"Page {i + 1}")
                            row_offset += 1
                            for line in text.split("\n"):
                                ws.cell(row=row_offset, column=1, value=line)
                                row_offset += 1
                            row_offset += 1

            wb.save(output_path)

            logger.info(f"PDF to Excel complete: {tables_found} tables extracted -> {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"PDF to Excel failed: {e}", exc_info=True)
            raise RuntimeError(f"PDF to Excel failed: {str(e)}")
